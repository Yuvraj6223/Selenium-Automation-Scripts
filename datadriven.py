from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
import time

workbook = load_workbook('trialidp.xlsx')
sheet = workbook.active

driver = webdriver.Chrome()
driver.maximize_window()

for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
    username = row[0]
    password = row[1]
    
    driver.get('https://saucedemo.com/')
    driver.find_element(By.ID, "user-name").send_keys(username)
    driver.find_element(By.ID, "password").send_keys(password)
    driver.find_element(By.ID, "login-button").click()
    
    time.sleep(3)  

driver.quit()
